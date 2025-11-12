import pandas as pd
import sys
from pathlib import Path
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from PyQt6.QtGui import QTextDocument
from PyQt6.QtPrintSupport import QPrinter
from PyQt6.QtCore import QSizeF
from logic.formatting import _normalize_period, _parse_number_safe
from logic.audit import run_basic_audit, compare_periods

# Import DataRepository pour PostgreSQL
sys.path.insert(0, str(Path(__file__).parent.parent))
from app.services.data_repo import DataRepository


def df_resume_mois(period):
    df = _load_period_data(period)
    if df.empty:
        return pd.DataFrame()

    montant_col = _find_column(df, ["Montant", "montant"])
    emp_col = _find_column(df, ["Matricule", "matricule"])

    if not montant_col:
        return pd.DataFrame()

    df["_montant_num"] = df[montant_col].apply(_parse_number_safe)

    net_total = df["_montant_num"].sum()
    deductions = df[df["_montant_num"] < 0]["_montant_num"].sum()
    brut_total = net_total + abs(deductions)
    nb_employes = df[emp_col].nunique() if emp_col else len(df)
    net_moyen = net_total / nb_employes if nb_employes > 0 else 0

    return pd.DataFrame(
        {
            "Indicateur": [
                "Net total",
                "Brut total",
                "Déductions",
                "Nb employés",
                "Net moyen",
            ],
            "Valeur": [net_total, brut_total, deductions, nb_employes, net_moyen],
        }
    )


def df_detail_employe(period):
    df = _load_period_data(period)
    if df.empty:
        return pd.DataFrame()

    emp_col = _find_column(df, ["Matricule", "matricule"])
    name_col = _find_column(df, ["Nom et prénom", "Nom", "nom"])
    cat_col = _find_column(df, ["Catégorie de paie", "categorie", "TypePaie"])
    montant_col = _find_column(df, ["Montant", "montant"])

    if not emp_col or not montant_col:
        return pd.DataFrame()

    df["_montant_num"] = df[montant_col].apply(_parse_number_safe)

    if cat_col:
        pivot = df.pivot_table(
            index=[emp_col, name_col] if name_col else [emp_col],
            columns=cat_col,
            values="_montant_num",
            aggfunc="sum",
            fill_value=0,
        ).reset_index()
    else:
        pivot = (
            df.groupby([emp_col, name_col] if name_col else [emp_col])
            .agg({"_montant_num": "sum"})
            .reset_index()
        )
        pivot.columns = [emp_col] + ([name_col] if name_col else []) + ["Total"]

    if "Gains" in pivot.columns:
        pivot["Gains"] = pivot["Gains"]
    for col in pivot.columns:
        if col.lower() in ["déductions", "deductions", "assurances"]:
            pivot[col] = -pivot[col].abs()

    if "Gains" in pivot.columns and any(
        c.lower() in ["déductions", "deductions"] for c in pivot.columns
    ):
        ded_col = next(
            (c for c in pivot.columns if c.lower() in ["déductions", "deductions"]),
            None,
        )
        if ded_col:
            pivot["Net"] = pivot["Gains"] + pivot[ded_col]

    return pivot


def df_rep_code_paie(period):
    df = _load_period_data(period)
    if df.empty:
        return pd.DataFrame()

    code_col = _find_column(df, ["Code de paie", "code_paie", "Code"])
    desc_col = _find_column(
        df, ["Description code de paie", "description", "Description"]
    )
    montant_col = _find_column(df, ["Montant", "montant"])

    if not code_col or not montant_col:
        return pd.DataFrame()

    df["_montant_num"] = df[montant_col].apply(_parse_number_safe)

    if desc_col:
        result = (
            df.groupby([code_col, desc_col]).agg({"_montant_num": "sum"}).reset_index()
        )
        result.columns = ["Code", "Description", "Montant"]
    else:
        result = df.groupby(code_col).agg({"_montant_num": "sum"}).reset_index()
        result.columns = ["Code", "Montant"]

    result = result.sort_values("Montant", ascending=False)

    return result


def df_rep_poste_budgetaire(period):
    df = _load_period_data(period)
    if df.empty:
        return pd.DataFrame()

    poste_col = _find_column(df, ["Poste budgétaire", "poste_budgetaire", "Poste"])
    montant_col = _find_column(df, ["Montant", "montant"])

    if not poste_col or not montant_col:
        return pd.DataFrame()

    df["_montant_num"] = df[montant_col].apply(_parse_number_safe)

    result = df.groupby(poste_col).agg({"_montant_num": "sum"}).reset_index()
    result.columns = ["Poste budgétaire", "Montant"]
    result = result.sort_values("Montant", ascending=False)

    return result


def df_evolution_12p(period):
    try:
        # PostgreSQL: charger toutes les données
        from config.config_manager import get_dsn

        dsn = get_dsn()
        repo = DataRepository(dsn, min_size=1, max_size=2)
        rows = repo.run_query(
            "SELECT * FROM payroll.imported_payroll_master", fetch_all=True
        )

        # Récupérer les noms de colonnes
        with repo.get_connection() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT * FROM payroll.imported_payroll_master LIMIT 0")
                columns = [desc[0] for desc in cur.description]
        repo.close()

        df = pd.DataFrame(rows, columns=columns)

        if df.empty:
            return pd.DataFrame()

        date_col = _find_column(df, ["Date de paie", "date_paie", "Date"])
        montant_col = _find_column(df, ["Montant", "montant"])

        if not date_col or not montant_col:
            return pd.DataFrame()

        df["_period"] = (
            pd.to_datetime(df[date_col], errors="coerce").dt.to_period("M").astype(str)
        )
        df["_montant_num"] = df[montant_col].apply(_parse_number_safe)

        periods = df["_period"].dropna().unique()
        periods = sorted([p for p in periods if p != "NaT"])[-12:]

        result = (
            df[df["_period"].isin(periods)]
            .groupby("_period")
            .agg({"_montant_num": "sum"})
            .reset_index()
        )
        result.columns = ["Période", "Net total"]

        return result
    except Exception as _exc:
        return pd.DataFrame()


def df_anomalies(period):
    audit = run_basic_audit(period)
    anomalies_df = audit.get("anomalies_df", pd.DataFrame())
    return anomalies_df


def df_comparaison(p1, p2):
    comp = compare_periods(p1, p2)

    data = {
        "Indicateur": ["Delta net", "Variation %"],
        "Valeur": [comp.get("delta_net", 0), comp.get("pct", 0)],
    }

    df = pd.DataFrame(data)

    return df


def export_excel_resume(period, filepath):
    df = df_resume_mois(period)
    _export_excel_generic(df, filepath, f"Résumé - {period}", period)


def export_excel_detail_employe(period, filepath):
    df = df_detail_employe(period)
    _export_excel_generic(df, filepath, f"Détail employés - {period}", period)


def export_excel_rep_code_paie(period, filepath):
    df = df_rep_code_paie(period)
    _export_excel_generic(df, filepath, f"Répartition codes - {period}", period)


def export_excel_rep_poste(period, filepath):
    df = df_rep_poste_budgetaire(period)
    _export_excel_generic(df, filepath, f"Répartition postes - {period}", period)


def export_excel_evolution(period, filepath):
    df = df_evolution_12p(period)
    _export_excel_generic(df, filepath, "Évolution 12 périodes", period)


def export_excel_anomalies(period, filepath):
    df = df_anomalies(period)
    _export_excel_generic(df, filepath, f"Anomalies - {period}", period)


def export_excel_comparaison(p1, p2, filepath):
    df = df_comparaison(p1, p2)
    _export_excel_generic(df, filepath, f"Comparaison {p1} vs {p2}", f"{p1} vs {p2}")


def _export_excel_generic(df, filepath, title, subtitle):
    if df.empty:
        df = pd.DataFrame({"Info": ["Aucune donnée disponible"]})

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Rapport", index=False, startrow=3)

        wb = writer.book
        ws = writer.sheets["Rapport"]

        ws.merge_cells(
            start_row=1, start_column=1, end_row=1, end_column=len(df.columns)
        )
        title_cell = ws.cell(1, 1)
        title_cell.value = title
        title_cell.font = Font(size=16, bold=True, color="1e3a8a")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells(
            start_row=2, start_column=1, end_row=2, end_column=len(df.columns)
        )
        subtitle_cell = ws.cell(2, 1)
        subtitle_cell.value = f"Période: {subtitle} | Généré le {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        subtitle_cell.font = Font(size=10, italic=True)
        subtitle_cell.alignment = Alignment(horizontal="center")

        header_fill = PatternFill(
            start_color="d1d5db", end_color="d1d5db", fill_type="solid"
        )
        header_font = Font(bold=True)
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(4, col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

            max_length = max(
                len(str(col_name)),
                df[col_name].astype(str).str.len().max() if not df.empty else 0,
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(
                max_length + 2, 50
            )

        for row_idx in range(5, 5 + len(df)):
            for col_idx in range(1, len(df.columns) + 1):
                cell = ws.cell(row_idx, col_idx)
                cell.border = border

                if (
                    "montant" in df.columns[col_idx - 1].lower()
                    or "valeur" in df.columns[col_idx - 1].lower()
                ):
                    cell.number_format = "#,##0.00"


def export_pdf_resume(period, filepath):
    df = df_resume_mois(period)
    _export_pdf_generic(df, filepath, f"Résumé - {period}", period)


def export_pdf_detail_employe(period, filepath):
    df = df_detail_employe(period)
    _export_pdf_generic(df, filepath, f"Détail employés - {period}", period)


def export_pdf_rep_code_paie(period, filepath):
    df = df_rep_code_paie(period)
    _export_pdf_generic(df, filepath, f"Répartition codes - {period}", period)


def export_pdf_rep_poste(period, filepath):
    df = df_rep_poste_budgetaire(period)
    _export_pdf_generic(df, filepath, f"Répartition postes - {period}", period)


def export_pdf_evolution(period, filepath):
    df = df_evolution_12p(period)
    _export_pdf_generic(df, filepath, "Évolution 12 périodes", period)


def export_pdf_anomalies(period, filepath):
    df = df_anomalies(period)
    _export_pdf_generic(df, filepath, f"Anomalies - {period}", period)


def export_pdf_comparaison(p1, p2, filepath):
    df = df_comparaison(p1, p2)
    _export_pdf_generic(df, filepath, f"Comparaison {p1} vs {p2}", f"{p1} vs {p2}")


def _export_pdf_generic(df, filepath, title, subtitle):
    if df.empty:
        df = pd.DataFrame({"Info": ["Aucune donnée disponible"]})

    html = f"""
    <html>
    <head>
        <style>
            body {{ font-family: Arial, sans-serif; margin: 20px; }}
            h1 {{ color: #1e3a8a; text-align: center; }}
            h2 {{ color: #4b5563; text-align: center; font-size: 12px; }}
            table {{ border-collapse: collapse; width: 100%; margin-top: 20px; }}
            th {{ background-color: #d1d5db; font-weight: bold; padding: 8px; border: 1px solid #9ca3af; text-align: left; }}
            td {{ padding: 8px; border: 1px solid #d1d5db; }}
            tr:nth-child(even) {{ background-color: #f9fafb; }}
            .footer {{ position: fixed; bottom: 0; width: 100%; text-align: center; font-size: 10px; color: #6b7280; }}
        </style>
    </head>
    <body>
        <h1>{title}</h1>
        <h2>Période: {subtitle} | Généré le {datetime.now().strftime('%Y-%m-%d %H:%M')}</h2>
        <table>
            <thead>
                <tr>
    """

    for col in df.columns:
        html += f"<th>{col}</th>"

    html += """
                </tr>
            </thead>
            <tbody>
    """

    for _, row in df.iterrows():
        html += "<tr>"
        for val in row:
            if isinstance(val, (int, float)):
                html += f"<td style='text-align: right;'>{val:,.2f}</td>"
            else:
                html += f"<td>{val}</td>"
        html += "</tr>"

    html += """
            </tbody>
        </table>
        <div class="footer">
            <p>Système de Contrôle de la Paie (SCP) - Page {page}</p>
        </div>
    </body>
    </html>
    """

    doc = QTextDocument()
    doc.setHtml(html)
    doc.setPageSize(QSizeF(210, 297))

    printer = QPrinter(QPrinter.PrinterMode.HighResolution)
    printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
    printer.setOutputFileName(filepath)

    # PyQt6: setPageOrientation au lieu de setOrientation
    from PyQt6.QtGui import QPageLayout

    if len(df.columns) > 6:
        printer.setPageOrientation(QPageLayout.Orientation.Landscape)
    else:
        printer.setPageOrientation(QPageLayout.Orientation.Portrait)

    doc.print(printer)  # PyQt6: print au lieu de print_


def _load_period_data(period):
    """Charge les données de période depuis PostgreSQL."""
    try:
        from config.config_manager import get_dsn

        dsn = get_dsn()
        repo = DataRepository(dsn, min_size=1, max_size=2)

        if period:
            period_str = (
                _normalize_period(period).strftime("%Y-%m")
                if hasattr(period, "strftime")
                else str(period)
            )
            query = """
                SELECT * FROM payroll.imported_payroll_master 
                WHERE "date de paie " LIKE %s || '%%'
            """
            rows = repo.run_query(query, (period_str,), fetch_all=True)
        else:
            query = "SELECT * FROM payroll.imported_payroll_master"
            rows = repo.run_query(query, fetch_all=True)

        # Récupérer les noms de colonnes
        with repo.get_connection() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT * FROM payroll.imported_payroll_master LIMIT 0")
                columns = [desc[0] for desc in cur.description]
        repo.close()

        return pd.DataFrame(rows, columns=columns)
    except Exception as e:
        print(f"Erreur chargement données: {e}")
        return pd.DataFrame()


def _find_column(df, candidates):
    for c in candidates:
        if c in df.columns:
            return c
        for col in df.columns:
            if col.lower() == c.lower():
                return col
    return None
