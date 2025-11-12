import argparse
from pathlib import Path
from collections import defaultdict
from datetime import datetime, date
import re

try:
    from openpyxl import load_workbook
except Exception as e:
    raise SystemExit(
        "openpyxl n'est pas installé. Installez-le avec: py -m pip install openpyxl"
    )


def normalize_header(name: str) -> str:
    if name is None:
        return ""
    s = str(name).strip().lower()
    s = re.sub(r"\s+", "_", s)
    s = (
        s.replace("é", "e")
        .replace("è", "e")
        .replace("ê", "e")
        .replace("à", "a")
        .replace("ù", "u")
        .replace("ô", "o")
    )
    return s


def parse_date(value) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, date):
        return value
    if isinstance(value, datetime):
        return value.date()
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    return None


def to_number(value) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace(" ", "").replace(",", ".")
    try:
        return float(s)
    except Exception:
        return 0.0


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--file", required=True, help="Chemin vers le fichier .xlsx")
    p.add_argument(
        "--sheet", default=None, help="Nom de la feuille (par défaut: 1ere feuille)"
    )
    args = p.parse_args()

    xlsx = Path(args.file)
    if not xlsx.exists():
        raise SystemExit(f"Fichier introuvable: {xlsx}")

    wb = load_workbook(filename=str(xlsx), data_only=True)
    ws = (
        wb[args.sheet]
        if args.sheet and args.sheet in wb.sheetnames
        else wb[wb.sheetnames[0]]
    )

    # Lire l'entête
    headers = []
    for cell in ws[1]:
        headers.append(normalize_header(cell.value))

    header_to_idx = {h: i for i, h in enumerate(headers)}

    # Détecter colonnes
    date_candidates = ["date_paie", "date", "pay_date", "date_de_paie", "datepaie"]
    categorie_candidates = [
        "categorie_paie",
        "categorie_de_paie",
        "categorie",
        "category",
    ]
    amount_unit_candidates = [
        "montant",
        "montant_employe",
        "amount",
        "valeur",
        "montant_net",
    ]
    amount_cents_candidates = ["montant_cents", "amount_cents"]

    def find_col(cands):
        for c in cands:
            if c in header_to_idx:
                return c
        return None

    date_col = find_col(date_candidates)
    cat_col = find_col(categorie_candidates)
    amt_col = find_col(amount_unit_candidates)
    amtc_col = find_col(amount_cents_candidates)

    if not date_col:
        raise SystemExit(f"Colonne date introuvable. Entêtes normalisées: {headers}")
    if not cat_col:
        raise SystemExit(
            f"Colonne categorie_paie introuvable. Entêtes normalisées: {headers}"
        )
    if not (amt_col or amtc_col):
        raise SystemExit(
            f"Colonne montant OU montant_cents introuvable. Entêtes: {headers}"
        )

    sums_by_date: dict[str, float] = defaultdict(float)

    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_date = (
            row[header_to_idx[date_col]]
            if header_to_idx.get(date_col) is not None
            else None
        )
        d = parse_date(raw_date)
        if not d:
            continue

        raw_cat = (
            row[header_to_idx[cat_col]]
            if header_to_idx.get(cat_col) is not None
            else None
        )
        cat_norm = normalize_header(raw_cat)
        if cat_norm != "gains":
            continue

        amount = 0.0
        if amt_col:
            raw_amt = row[header_to_idx[amt_col]]
            amount = to_number(raw_amt)
        else:
            raw_cents = row[header_to_idx[amtc_col]]
            amount = to_number(raw_cents) / 100.0

        key = d.isoformat()
        sums_by_date[key] += amount

    # Sort by date
    items = sorted(sums_by_date.items(), key=lambda x: x[0])

    out_dir = Path("logs/analytics")
    out_dir.mkdir(parents=True, exist_ok=True)
    out_file = out_dir / "masse_from_file.csv"
    with out_file.open("w", encoding="utf-8") as f:
        f.write("date_paie;masse_salariale\n")
        for k, v in items:
            f.write(f"{k};{v}\n")

    print("Masse salariale (Gains uniquement, depuis le fichier):")
    for k, v in items:
        print(f"{k};{v}")
    print(f"\nExport: {out_file.resolve()}")


if __name__ == "__main__":
    main()
