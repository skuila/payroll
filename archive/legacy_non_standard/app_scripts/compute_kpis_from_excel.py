import argparse
from pathlib import Path
from collections import defaultdict, Counter
from datetime import datetime, date
import re

try:
    from openpyxl import load_workbook
except Exception:
    raise SystemExit("openpyxl non installé. Installez-le: py -m pip install openpyxl")


def normalize_header(name: str) -> str:
    if name is None:
        return ""
    s = str(name).strip().lower()
    s = (
        s.replace("é", "e")
        .replace("è", "e")
        .replace("ê", "e")
        .replace("à", "a")
        .replace("ù", "u")
        .replace("ô", "o")
    )
    s = re.sub(r"\s+|/|\\|\-", "_", s)
    return s


def parse_date(value) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, date):
        return value
    if isinstance(value, datetime):
        return value.date()
    s = str(value).strip()
    for fmt in (
        "%Y-%m-%d",
        "%Y-%m-%dT%H:%M:%S",
        "%d/%m/%Y",
        "%m/%d/%Y",
        "%d-%m-%Y",
        "%Y/%m/%d",
    ):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    return None


def to_number(value) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace(" ", "").replace(",", ".")
    try:
        return float(s)
    except Exception:
        return 0.0


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--file", required=True, help="Chemin du fichier .xlsx")
    p.add_argument("--sheet", default=None, help="Nom de la feuille (par défaut 1ère)")
    args = p.parse_args()

    xlsx = Path(args.file)
    if not xlsx.exists():
        raise SystemExit(f"Fichier introuvable: {xlsx}")

    wb = load_workbook(filename=str(xlsx), data_only=True)
    ws = (
        wb[args.sheet]
        if args.sheet and args.sheet in wb.sheetnames
        else wb[wb.sheetnames[0]]
    )

    # Entêtes
    headers = [normalize_header(c.value) for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}

    # Candidats pour colonnes
    date_cols = ["date_paie", "date_de_paie", "date", "pay_date", "datepaie"]
    cat_cols = ["categorie_paie", "categorie_de_paie", "categorie", "category"]
    code_cols = ["code_paie", "code_de_paie", "code"]
    lib_cols = ["description_code_paie", "desc_code_de_paie", "libelle_paie", "libelle"]
    poste_cols = ["poste_budgetaire", "poste", "postebudgetaire"]
    poste_lib_cols = [
        "description_poste_budgetaire",
        "desc_poste_budgetaire",
        "libelle_poste",
    ]
    emp_cols = ["nom_employe", "employe", "nom_prenom", "employe_nom"]
    mat_cols = ["matricule", "id_employe", "employee_id"]
    montant_cols = ["montant", "montant_employe", "amount", "montant_net"]
    montant_cents_cols = ["montant_cents", "amount_cents"]
    part_cols = ["part_employeur", "partemployeur", "employer_share"]
    part_cents_cols = ["part_employeur_cents", "employer_share_cents"]

    def pick(cands):
        for c in cands:
            if c in idx:
                return c
        return None

    date_col = pick(date_cols)
    cat_col = pick(cat_cols)
    code_col = pick(code_cols)
    lib_col = pick(lib_cols)
    poste_col = pick(poste_cols)
    poste_lib_col = pick(poste_lib_cols)
    emp_col = pick(emp_cols)
    mat_col = pick(mat_cols)
    montant_col = pick(montant_cols)
    montant_cents_col = pick(montant_cents_cols)
    part_col = pick(part_cols)
    part_cents_col = pick(part_cents_cols)

    if not date_col:
        raise SystemExit(f"Colonne date introuvable. Entêtes: {headers}")
    if not cat_col:
        raise SystemExit(f"Colonne categorie de paie introuvable. Entêtes: {headers}")
    if not (montant_col or montant_cents_col):
        raise SystemExit(
            f"Colonne montant ou montant_cents introuvable. Entêtes: {headers}"
        )

    # Agrégations
    per_date = defaultdict(
        lambda: {"gains": 0.0, "deductions": 0.0, "part": 0.0, "net_sum": 0.0}
    )
    per_cat = defaultdict(lambda: {"part": 0.0, "count": 0})
    per_poste = defaultdict(lambda: {"gains": 0.0, "deductions": 0.0, "part": 0.0})
    per_emp = defaultdict(
        lambda: {"gains": 0.0, "deductions": 0.0, "part": 0.0, "count": 0}
    )
    per_code = defaultdict(lambda: {"cat": "", "part": 0.0, "count": 0})

    for row in ws.iter_rows(min_row=2, values_only=True):
        d = parse_date(row[idx[date_col]]) if date_col in idx else None
        if not d:
            continue
        key_date = d.isoformat()

        cat = normalize_header(row[idx[cat_col]]) if cat_col in idx else ""
        code = (
            str(row[idx[code_col]]).strip()
            if code_col in idx and row[idx[code_col]] is not None
            else ""
        )
        lib = (
            str(row[idx[lib_col]]).strip()
            if lib_col in idx and row[idx[lib_col]] is not None
            else ""
        )
        poste = (
            str(row[idx[poste_col]]).strip()
            if poste_col in idx and row[idx[poste_col]] is not None
            else ""
        )
        poste_lib = (
            str(row[idx[poste_lib_col]]).strip()
            if poste_lib_col in idx and row[idx[poste_lib_col]] is not None
            else ""
        )
        emp = (
            str(row[idx[emp_col]]).strip()
            if emp_col in idx and row[idx[emp_col]] is not None
            else ""
        )
        mat = (
            str(row[idx[mat_col]]).strip()
            if mat_col in idx and row[idx[mat_col]] is not None
            else ""
        )

        if montant_col:
            montant = to_number(row[idx[montant_col]])
        else:
            montant = to_number(row[idx[montant_cents_col]]) / 100.0
        part = 0.0
        if part_col:
            part = to_number(row[idx[part_col]])
        elif part_cents_col:
            part = to_number(row[idx[part_cents_col]]) / 100.0

        # Arrondir à 2 décimales (comme Excel affiche)
        montant = round(montant, 2)
        part = round(part, 2)

        # Net = somme de tous les montants employés (catégorie ignorée)
        per_date[key_date]["net_sum"] += montant

        if cat == "gains":
            per_date[key_date]["gains"] += montant
            per_poste[(key_date, poste, poste_lib)]["gains"] += montant
            per_emp[(key_date, mat, emp)]["gains"] += montant
        elif (montant < 0) or cat.startswith("deduc"):
            per_date[key_date]["deductions"] += montant
            per_poste[(key_date, poste, poste_lib)]["deductions"] += montant
            per_emp[(key_date, mat, emp)]["deductions"] += montant

        per_date[key_date]["part"] += part
        per_poste[(key_date, poste, poste_lib)]["part"] += part
        per_emp[(key_date, mat, emp)]["part"] += part
        per_emp[(key_date, mat, emp)]["count"] += 1

        per_cat[(key_date, cat, code)]["part"] += part
        per_cat[(key_date, cat, code)]["count"] += 1
        per_code[(key_date, code)]["cat"] = cat
        per_code[(key_date, code)]["part"] += part
        per_code[(key_date, code)]["count"] += 1

    out_dir = Path("logs/analytics/excel_kpis")
    out_dir.mkdir(parents=True, exist_ok=True)

    # v_masse_salariale-like
    with (out_dir / "masse.csv").open("w", encoding="utf-8") as f:
        f.write("date_paie;gains;deductions;net;part_employeur;masse_salariale\n")
        for k in sorted(per_date.keys()):
            g = round(per_date[k]["gains"], 2)
            d_ = round(per_date[k]["deductions"], 2)
            p = round(per_date[k]["part"], 2)
            net = round(per_date[k]["net_sum"], 2)
            masse = round(g + p, 2)
            f.write(f"{k};{g};{d_};{net};{p};{masse}\n")

    # v_categories-like (pas de total_combine)
    with (out_dir / "categories.csv").open("w", encoding="utf-8") as f:
        f.write("date_paie;categorie_paie;code_paie;part_employeur;lignes\n")
        for (dte, cat, code), agg in sorted(per_cat.items()):
            f.write(f"{dte};{cat};{code};{agg['part']};{agg['count']}\n")

    # v_postes-like
    with (out_dir / "postes.csv").open("w", encoding="utf-8") as f:
        f.write(
            "date_paie;poste_budgetaire;description_poste_budgetaire;gains;deductions;part_employeur;masse_salariale\n"
        )
        for (dte, poste, poste_lib), agg in sorted(per_poste.items()):
            masse = agg["gains"] + agg["part"]
            f.write(
                f"{dte};{poste};{poste_lib};{agg['gains']};{agg['deductions']};{agg['part']};{masse}\n"
            )

    # v_employes-like
    with (out_dir / "employes.csv").open("w", encoding="utf-8") as f:
        f.write(
            "date_paie;matricule;nom_employe;gains;deductions;part_employeur;lignes;masse_salariale\n"
        )
        for (dte, mat, emp), agg in sorted(per_emp.items()):
            masse = agg["gains"] + agg["part"]
            f.write(
                f"{dte};{mat};{emp};{agg['gains']};{agg['deductions']};{agg['part']};{agg['count']};{masse}\n"
            )

    # codes (cat + part, pour contrôle)
    with (out_dir / "codes.csv").open("w", encoding="utf-8") as f:
        f.write("date_paie;code_paie;categorie_paie;part_employeur;lignes\n")
        for (dte, code), agg in sorted(per_code.items()):
            f.write(f"{dte};{code};{agg['cat']};{agg['part']};{agg['count']}\n")

    # Affichage synthèse
    print(
        "Masse salariale (gains + part_employeur) et Net (gains + deductions) par date:"
    )
    for k in sorted(per_date.keys()):
        g = round(per_date[k]["gains"], 2)
        p = round(per_date[k]["part"], 2)
        d_ = round(per_date[k]["deductions"], 2)
        net = round(per_date[k]["net_sum"], 2)
        masse = round(g + p, 2)
        print(
            f"{k};gains={g};deductions={d_};net={net};part_employeur={p};masse={masse}"
        )
    print(f"\nExports dans: {out_dir.resolve()}")


if __name__ == "__main__":
    main()
