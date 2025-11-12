import argparse
from pathlib import Path
import re
from datetime import datetime, date
from typing import Optional

try:
    import psycopg
except Exception:
    raise SystemExit(
        "psycopg non installé. Installez-le: py -m pip install psycopg[binary]"
    )

try:
    from openpyxl import load_workbook
except Exception:
    raise SystemExit("openpyxl non installé. Installez-le: py -m pip install openpyxl")


def normalize_header(name: str) -> str:
    if name is None:
        return ""
    s = str(name).strip().lower()
    s = (
        s.replace("é", "e")
        .replace("è", "e")
        .replace("ê", "e")
        .replace("à", "a")
        .replace("ù", "u")
        .replace("ô", "o")
    )
    s = re.sub(r"\s+|/|\\|\-", "_", s)
    return s


def parse_date(value) -> Optional[date]:
    if value is None or value == "":
        return None
    if isinstance(value, date):
        return value
    if isinstance(value, datetime):
        return value.date()
    s = str(value).strip()
    for fmt in (
        "%Y-%m-%d",
        "%Y-%m-%dT%H:%M:%S",
        "%d/%m/%Y",
        "%m/%d/%Y",
        "%d-%m-%Y",
        "%Y/%m/%d",
    ):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    return None


def to_number(value) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace(" ", "").replace(",", ".")
    try:
        return float(s)
    except Exception:
        return 0.0


def main():
    p = argparse.ArgumentParser()
    p.add_argument(
        "--dsn",
        required=True,
        help="DSN PostgreSQL ex: postgresql://user:pass@host:port/db",
    )
    p.add_argument("--file", required=True, help="Chemin du fichier .xlsx")
    p.add_argument("--sheet", default=None, help="Nom de la feuille (par défaut 1ère)")
    p.add_argument(
        "--truncate",
        action="store_true",
        help="Vider paie.stg_paie_transactions avant import",
    )
    args = p.parse_args()

    xlsx = Path(args.file)
    if not xlsx.exists():
        raise SystemExit(f"Fichier introuvable: {xlsx}")

    wb = load_workbook(filename=str(xlsx), data_only=True)
    ws = (
        wb[args.sheet]
        if args.sheet and args.sheet in wb.sheetnames
        else wb[wb.sheetnames[0]]
    )

    headers = [normalize_header(c.value) for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}

    # Mapping colonnes Excel → staging
    col_map = {
        "date_paie": next(
            (c for c in ("date_paie", "date_de_paie", "date", "pay_date") if c in idx),
            None,
        ),
        "matricule": next((c for c in ("matricule", "id_employe") if c in idx), None),
        "nom_prenom": next(
            (c for c in ("nom_employe", "employe", "nom_prenom") if c in idx), None
        ),
        "code_emploi": next((c for c in ("code_emploi",) if c in idx), None),
        "titre_emploi": next(
            (c for c in ("titre_emploi", "titre_d_emploi") if c in idx), None
        ),
        "categorie_emploi": next(
            (c for c in ("categorie_emploi", "categorie_d_emploi") if c in idx), None
        ),
        "code_paie": next(
            (c for c in ("code_paie", "code_de_paie", "code") if c in idx), None
        ),
        "libelle_paie": next(
            (
                c
                for c in (
                    "description_code_paie",
                    "desc_code_de_paie",
                    "libelle_paie",
                    "libelle",
                )
                if c in idx
            ),
            None,
        ),
        "poste_budgetaire": next(
            (c for c in ("poste_budgetaire", "poste") if c in idx), None
        ),
        "libelle_poste": next(
            (
                c
                for c in (
                    "description_poste_budgetaire",
                    "desc_poste_budgetaire",
                    "libelle_poste",
                )
                if c in idx
            ),
            None,
        ),
        "categorie_paie": next(
            (
                c
                for c in ("categorie_paie", "categorie_de_paie", "categorie")
                if c in idx
            ),
            None,
        ),
        "montant": next(
            (c for c in ("montant", "montant_employe", "amount") if c in idx), None
        ),
        "montant_cents": next((c for c in ("montant_cents",) if c in idx), None),
        "part_employeur": next(
            (c for c in ("part_employeur", "employer_share") if c in idx), None
        ),
        "part_employeur_cents": next(
            (c for c in ("part_employeur_cents",) if c in idx), None
        ),
        "montant_combine": next(
            (
                c
                for c in ("montant_combine", "mnt_cmb", "mnt_cmb", "mnt/cmb", "mntcmb")
                if c in idx
            ),
            None,
        ),
    }

    if not col_map["date_paie"]:
        raise SystemExit(f"Colonne date de paie introuvable. Entêtes: {headers}")

    batch_id = f"excel_import_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}"
    source_file = xlsx.name
    rows = []
    rownum = 1
    for r in ws.iter_rows(min_row=2, values_only=True):
        rownum += 1
        d = parse_date(r[idx[col_map["date_paie"]]]) if col_map["date_paie"] else None
        if not d:
            continue

        def get(col_key):
            col_name = col_map.get(col_key)
            return r[idx[col_name]] if col_name and col_name in idx else None

        montant = None
        if col_map["montant_cents"]:
            montant = to_number(get("montant_cents"))
        elif col_map["montant"]:
            montant = to_number(get("montant")) * 100.0
        else:
            montant = 0.0

        part = None
        if col_map["part_employeur_cents"]:
            part = to_number(get("part_employeur_cents"))
        elif col_map["part_employeur"]:
            part = to_number(get("part_employeur")) * 100.0
        else:
            part = 0.0

        combine_val = None
        if col_map["montant_combine"]:
            try:
                combine_val = to_number(get("montant_combine"))
            except Exception:
                combine_val = None

        rows.append(
            {
                "source_batch_id": batch_id,
                "source_file": source_file,
                "source_row_number": rownum,
                "date_paie_raw": (get("date_paie") if col_map["date_paie"] else None),
                "matricule_raw": (get("matricule") if col_map["matricule"] else None),
                "nom_prenom_raw": (
                    get("nom_prenom") if col_map["nom_prenom"] else None
                ),
                "code_emploi_raw": (
                    get("code_emploi") if col_map["code_emploi"] else None
                ),
                "titre_emploi_raw": (
                    get("titre_emploi") if col_map["titre_emploi"] else None
                ),
                "categorie_emploi_raw": (
                    get("categorie_emploi") if col_map["categorie_emploi"] else None
                ),
                "code_paie_raw": (get("code_paie") if col_map["code_paie"] else None),
                "libelle_paie_raw": (
                    get("libelle_paie") if col_map["libelle_paie"] else None
                ),
                "poste_budgetaire_raw": (
                    get("poste_budgetaire") if col_map["poste_budgetaire"] else None
                ),
                "libelle_poste_raw": (
                    get("libelle_poste") if col_map["libelle_poste"] else None
                ),
                "montant_raw": (
                    get("montant")
                    if col_map["montant"]
                    else (get("montant_cents") if col_map["montant_cents"] else None)
                ),
                "part_employeur_raw": (
                    get("part_employeur")
                    if col_map["part_employeur"]
                    else (
                        get("part_employeur_cents")
                        if col_map["part_employeur_cents"]
                        else None
                    )
                ),
                "date_paie": d,
                "matricule": (get("matricule") if col_map["matricule"] else None),
                "nom_prenom": (get("nom_prenom") if col_map["nom_prenom"] else None),
                "code_emploi": (get("code_emploi") if col_map["code_emploi"] else None),
                "titre_emploi": (
                    get("titre_emploi") if col_map["titre_emploi"] else None
                ),
                "categorie_emploi": (
                    get("categorie_emploi") if col_map["categorie_emploi"] else None
                ),
                "code_paie": (get("code_paie") if col_map["code_paie"] else None),
                "libelle_paie": (
                    get("libelle_paie") if col_map["libelle_paie"] else None
                ),
                "poste_budgetaire": (
                    get("poste_budgetaire") if col_map["poste_budgetaire"] else None
                ),
                "libelle_poste": (
                    get("libelle_poste") if col_map["libelle_poste"] else None
                ),
                "categorie_paie": (
                    get("categorie_paie") if col_map["categorie_paie"] else None
                ),
                "montant_cents": int(round(montant or 0.0)),
                "part_employeur_cents": int(round(part or 0.0)),
                "montant_combine": combine_val,
            }
        )

    with psycopg.connect(args.dsn) as conn:
        conn.execute("SET TIME ZONE 'UTC'")
        with conn.transaction():
            if args.truncate:
                conn.execute(
                    "TRUNCATE TABLE paie.stg_paie_transactions RESTART IDENTITY"
                )
            # Insert batched
            sql = (
                "INSERT INTO paie.stg_paie_transactions ("
                "source_batch_id, source_file, source_row_number,"
                "date_paie_raw, matricule_raw, nom_prenom_raw, code_emploi_raw, titre_emploi_raw, categorie_emploi_raw, code_paie_raw, libelle_paie_raw, poste_budgetaire_raw, libelle_poste_raw, montant_raw, part_employeur_raw,"
                "date_paie, matricule, nom_prenom, code_emploi, titre_emploi, categorie_emploi, code_paie, libelle_paie, poste_budgetaire, libelle_poste, categorie_paie,"
                "montant_cents, part_employeur_cents, is_valid"
                ") VALUES ("
                "%(source_batch_id)s, %(source_file)s, %(source_row_number)s,"
                "%(date_paie_raw)s, %(matricule_raw)s, %(nom_prenom_raw)s, %(code_emploi_raw)s, %(titre_emploi_raw)s, %(categorie_emploi_raw)s, %(code_paie_raw)s, %(libelle_paie_raw)s, %(poste_budgetaire_raw)s, %(libelle_poste_raw)s, %(montant_raw)s, %(part_employeur_raw)s,"
                "%(date_paie)s, %(matricule)s, %(nom_prenom)s, %(code_emploi)s, %(titre_emploi)s, %(categorie_emploi)s, %(code_paie)s, %(libelle_paie)s, %(poste_budgetaire)s, %(libelle_poste)s, %(categorie_paie)s,"
                "%(montant_cents)s, %(part_employeur_cents)s, true"
                ")"
            )
            # Chunk insert
            batch = 1000
            for i in range(0, len(rows), batch):
                with conn.cursor() as cur:
                    cur.executemany(sql, rows[i : i + batch])

    print(f"Import termine. Lignes insérées: {len(rows)}")


if __name__ == "__main__":
    main()
